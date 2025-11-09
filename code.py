from pathlib import Path
import csv
import datetime as dt
import fnmatch
import openpyxl

def _cell_value(ws, coord):
    c = ws[coord]
    v = c.value
    if v is None:
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                v = ws.cell(row=rng.min_row, column=rng.min_col).value
                break
    if isinstance(v, (dt.date, dt.datetime)):
        return v.isoformat()
    return v

def _ordered_union(*dicts):
    seen = set()
    res = []
    for d in dicts:
        for v in d.values():
            if v not in seen:
                seen.add(v)
                res.append(v)
    return res

def _filter_sheets(all_names, include=None, exclude=None):
    def match_any(name, patterns):
        return any(fnmatch.fnmatch(name, p) for p in patterns)
    names = list(all_names)
    if include:
        names = [n for n in names if match_any(n, include)]
    if exclude:
        names = [n for n in names if not match_any(n, exclude)]
    return names

EXCEL_PATH = "YourBoardingPassDotAero-2017-01-01.xlsx"
OUTPUT_CSV = "YourBoardingPassDotAero-2017-01-01.csv"

# Режим записи: "rows_per_sheet" — по строке на лист; "single_row_merged" — одна строка
MODE = "rows_per_sheet"

# Фильтр листов
SHEET_INCLUDE = []          
SHEET_EXCLUDE = []          

# Базовое сопоставление применится к каждому листу
MAPPING_DEFAULT = {
    "B3": "Name",
    "A5": "FlightNumber",
    "D7": "DepartureCode",
    "D5": "DepartureCity",
    "H7": "ArrivalCode",
    "H5": "ArrivalCity",
    "A3": "Sex",
    "F3": "Loyality",
    "A9": "Date",
    "C9": "Time",
    "E9": "CodeShare",
    "B7": "Gate",
    "E13": "TicketNumber",
    "H11": "Seat",
    "B13": "Fare",
    "H3": "Class"
}

# Переопределения/дополнения для конкретных листов
MAPPING_OVERRIDES = {}

def extract_from_workbook_many(xlsx_path, mapping_default, mapping_overrides=None,
                               include=None, exclude=None, mode="rows_per_sheet"):
    mapping_overrides = mapping_overrides or {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    sheets = _filter_sheets(wb.sheetnames, include, exclude)

    per_sheet_mapping = {}
    for s in sheets:
        m = dict(mapping_default)
        if s in mapping_overrides:
            m.update(mapping_overrides[s])
        if m:
            per_sheet_mapping[s] = m

    if not per_sheet_mapping:
        return [], []

    cols = _ordered_union(mapping_default, *[mapping_overrides.get(s, {}) for s in sheets])
    if mode == "rows_per_sheet":
        header = ["_sheet"] + cols
        rows = []
        for s, cell2col in per_sheet_mapping.items():
            row = {k: None for k in header}
            row["_sheet"] = s
            ws = wb[s]
            for coord, col in cell2col.items():
                try:
                    row[col] = _cell_value(ws, coord)
                except Exception:
                    row[col] = None
            rows.append(row)
        return header, rows

    if mode == "single_row_merged":
        header = cols
        row = {k: None for k in header}
        for s, cell2col in per_sheet_mapping.items():
            ws = wb[s]
            for coord, col in cell2col.items():
                if row[col] not in (None, ""):
                    continue
                try:
                    v = _cell_value(ws, coord)
                except Exception:
                    v = None
                if v not in (None, ""):
                    row[col] = v
        return header, [row]

    raise ValueError("Unknown MODE")

def write_csv(path, header, rows):
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in header})

header, data_rows = extract_from_workbook_many(
    EXCEL_PATH,
    MAPPING_DEFAULT,
    mapping_overrides=MAPPING_OVERRIDES,
    include=SHEET_INCLUDE,
    exclude=SHEET_EXCLUDE,
    mode=MODE,
)

write_csv(OUTPUT_CSV, header, data_rows)
print(f"Готово: {OUTPUT_CSV}, строк: {len(data_rows)}")